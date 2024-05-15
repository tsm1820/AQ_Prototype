"""
File: app\interfaces\db_connector.py

Place holder of Database connector object
"""
# Always set parent package, relative use case cause failure
import sys, os
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
sys.path.append(parent_dir)

from globalclass.osbasic import Fundamental as OSBASIC
import sqlite3 as SQlite
import openpyxl as PyExcel

# Constant path of database configuration file
DB_CONF_JSON = "/conf/db_conf.json"

class DB_Object():
    __db_obj = None

    def __init__(self, DB_Connector_Key):
        ## Configuration path << from conf .json >>
        self.__perform_opening(DB_Connector_Key)

    def __perform_opening(self, DB_Connector_Key):
        db_config = OSBASIC.loadConfiguration(DB_CONF_JSON)
        self.__db_obj = SQLite_Connector(db_config[DB_Connector_Key])
        pass

    def perform_sql(self, 
                    string_command:str ,
                    get_result:bool=False):
        ret_val = None

        self.__db_obj.perform_sql(string_command)

        if (get_result == True):
            ret_val = self.__db_obj.get_result()
            return ret_val

    def commit_update(self):
        self.__db_obj.commit_update()

class SQLite_Connector():
    __cursor = None
    __conn = None
    __result = None
    
    def __init__(self, 
                 path:str=""):
        self.__reconnect__(path)

    def __reconnect__(self, 
                      path:str):
        if (self.__conn == None):
            # This path is set to root folder
            self.__conn = SQlite.connect(path)
            self.__cursor = self.__conn.cursor()

    def get_result(self):
        return self.__result.fetchall()

    def perform_sql(self, 
                    string:str=""):
        if (self.__conn == None):
            self.__reconnect__()
        else:
            if len(string) != 0:
                self.__result = self.__cursor.execute(str(string))

    def commit_update(self):
        if (self.__conn != None):
            self.__conn.commit()

# Dumping the excel sheet to prepare the data
if __name__ == "__main__":

    def dump_traindata(DB_Obj):
        wb = PyExcel.load_workbook(OSBASIC.referFile("/model/traindata.xlsx"), read_only=True, data_only=True)
        ws = wb['train']
        string_1 = "INSERT INTO train_table(user_id, quiz_id, quiz_correct_ans, n_attempt)"

        data_list = list(ws.values) # Dump into worksheet # Faster method than extract each

        for i in range(1, len(data_list)): # Max row
            for j in range(0, len(data_list[i])): # Max column
                string_2 = "values(" + str(j+1) + "," + str(i) + "," + str(data_list[i][j]) + "," + "1)"
                DB_Obj.perform_sql(string_1 + string_2)
        
        DB_Obj.commit_update()

    def dump_celldata(DB_Obj):
        wb = PyExcel.load_workbook(OSBASIC.referFile("/model/traindata.xlsx"), read_only=True)
        ws = wb['cell']
        string_1 = "INSERT INTO mastery_table(user_id,"
        
        for i in range(1, (30 + 1)):
            if (i != 30):
                string_1 += "topic_" + str(i) + ","
            else:
                string_1 += "topic_" + str(i) + ")"

        if (ws.cell(ws.max_row, ws.max_column).value == None):
            max_col = ws.max_column - 1
        else:
            max_col = ws.max_column

        for i in range(1, max_col + 1):
            string_2 = "values(" + str(i) + ","
            for j in range (2, ws.max_row + 1):
                if (j != ws.max_row):
                    string_2 += str(ws.cell(j, i).value).strip() + ","
                else:
                    string_2 += str(ws.cell(j, i).value).strip() + ")"

            DB_Obj.perform_sql(string_1 + string_2)

        DB_Obj.commit_update()


    def dump_testdata(DB_Obj):
        wb = PyExcel.load_workbook(OSBASIC.referFile("/model/testdata.xlsx"), read_only=True)
        ws = wb['pretest_data']

        data_list = list(ws.values)

        data_list.pop(0) # Pop header
        string_1 = "INSERT INTO pretest_table(user_id, learner_response, cell_index)"

        for i in range(0, len(data_list)): # row walking // cell_index
            for j in range(1, len(data_list[i])): # column walking // User id
                string_2 = "values(" + str(j) + "," + str(data_list[i][j]) + "," + str(data_list[i][0]) + ")"
                DB_Obj.perform_sql(string_1 + string_2)

        DB_Obj.commit_update()
        

    def dump_quizpool(DB_Obj):
        """ Use this to dump the data from excel sheet"""
        cell_index_col = 1
        question_text_col = 3
        opt_1_col = 4
        opt_2_col = 5
        opt_3_col = 6
        opt_4_col = 7
        answer_col = 8
        mod_col = 9
        parent_col = 10
        sub_col = 11
        id_name_col = 12
        wb = PyExcel.load_workbook(OSBASIC.referFile("/model/quizpool.xlsx"), read_only=True)
        # Lock only this sheet
        ws = wb['currentQuiz']

        string_1 = "INSERT INTO quiz_pool(cell_index, question_text, opt_1, opt_2, opt_3, opt_4, answer, mod, parent, sub, id_name)"
        
        for i in range(1, ws.max_row + 1):
            if (i != 1):
                string_2 = " values("
                string_2 += str(ws.cell(i, cell_index_col).value).strip() + ","
                string_2 += "\"" + str(ws.cell(i, question_text_col).value).replace("\"", "'") + "\"" + "," # String
                string_2 += "\"" + str(ws.cell(i, opt_1_col).value).replace("\"", "'") + "\"" + "," # String
                string_2 += "\"" + str(ws.cell(i, opt_2_col).value).replace("\"", "'") + "\"" + "," # String
                string_2 += "\"" + str(ws.cell(i, opt_3_col).value).replace("\"", "'") + "\"" + "," # String
                string_2 += "\"" + str(ws.cell(i, opt_4_col).value).replace("\"", "'") + "\"" + "," # String
                string_2 += str(ws.cell(i, answer_col).value).strip()  + ","
                string_2 += "\"" + str(ws.cell(i, mod_col).value).replace("\"", "'") + "\"" + "," # String
                string_2 += str(ws.cell(i, parent_col).value).strip() + ","
                string_2 += "\"" + (str(ws.cell(i, sub_col).value).replace("\"", "'") + "\"" + ",") if ws.cell(i, sub_col).value != None else "\"\","
                string_2 += "\"" + str(ws.cell(i, id_name_col).value).replace("\"", "'") + "\"" # String
                string_2 += ")"
            
                DB_Obj.perform_sql(string_1 + string_2)

        # Transaction committing, must be called everytime of inserting
        DB_Obj.commit_update()
            
    def dump_explanation():

        DB_Obj = DB_Object("ADQ_DB")
        string_1 = "UPDATE quiz_pool set explanation = \"Explanation "
        for i in range(1, (531 + 1)):
            string_2 = str(i) + "\" where id="+str(i) + ";"
            DB_Obj.perform_sql(string_1 + string_2)

        DB_Obj.commit_update()


    def dump_namedata():
        DB_Obj = DB_Object("USERINFO_DB")
        wb = PyExcel.load_workbook(OSBASIC.referFile("/model/testdata.xlsx"), read_only=True)
        ws = wb['name']

        data_list = list(ws.values)

        string_1 = "INSERT INTO user_info(user_id, name)"

        for i in range(0, len(data_list)): # row walking // cell_index
            string_2 = "values(" + str(i+1) + ",\"" + data_list[i][0] + "\")"
            DB_Obj.perform_sql(string_1 + string_2)

        DB_Obj.commit_update()


    def dump_cell_info():
        list_item = ["1st" ,"2nd" ,"3rd" ,"4th" ,"5th" \
                     ,"6th" ,"7th" ,"8th" ,"9th" ,"10th" \
                     ,"11th" ,"12th" ,"13th" ,"14th" ,"15th" \
                     ,"16th" ,"17th" ,"18th" ,"19th" ,"20th" \
                     ,"21st" ,"22nd" ,"23rd" ,"24th" ,"25th" \
                     ,"26th" ,"27th" ,"28th" ,"29th" ,"30th"]
        DB_Obj = DB_Object("ADQ_DB")
        string_1 = "UPDATE cell_index_description set cell_description = \"Module Item "
        for i in range(1, (30 + 1)):
            string_2 = list_item[i-1] + "\" where cell_index="+str(i) + ";"
            DB_Obj.perform_sql(string_1 + string_2)

        DB_Obj.commit_update()
